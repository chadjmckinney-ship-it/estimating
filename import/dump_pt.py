import openpyxl, csv
from pathlib import Path
p = r"C:\Users\Chad\Estimate_Projects\import\New Current Worksheet.xlsm"
wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
print("ALL SHEETS:")
for n in wb.sheetnames:
    print(" ", n)
ws = wb["04-PT Slab on Grade"]
out = Path(r"C:\Users\Chad\Estimate_Projects\import\04-PT-SOG.csv")
with out.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    for i, row in enumerate(ws.iter_rows(max_row=90, max_col=36, values_only=True), 1):
        w.writerow(list(row))
print("WROTE", out)
# info job name
info = wb["Information"]
print("INFO rows:")
for i, row in enumerate(info.iter_rows(max_row=10, max_col=12, values_only=True), 1):
    print(i, row[:12])
wb.close()
