import openpyxl
p = r"C:\Users\Chad\Estimate_Projects\import\New Current Worksheet.xlsm"
wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
print("SHEETS", wb.sheetnames[:12], "...", len(wb.sheetnames))
for name in wb.sheetnames:
    if "Mono" in name or name in ("Information", "Summary", "SLAB ON GRADE"):
        ws = wb[name]
        rows = []
        for i, row in enumerate(ws.iter_rows(max_row=12, max_col=8, values_only=True), 1):
            rows.append(row)
        print("---", name, "---")
        for r in rows[:6]:
            print(r[:8])
wb.close()
